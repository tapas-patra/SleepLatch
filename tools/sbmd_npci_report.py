#!/usr/bin/env python3

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

IST = "Asia/Kolkata"

THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")


@dataclass
class ParsedPayload:
    field_name: str
    request_line: str
    body: str
    root_tag: str
    flat: Dict[str, str]
    normalized_flat: Dict[str, str]


@dataclass
class Exchange:
    sequence_index: int
    case_name: str
    case_type: str
    source_file: str
    timestamp_utc: str
    timestamp_ist: str
    api_flow: str
    api_dir: str
    api_target: str
    caller: str
    level: str
    log_message: str
    request_payload: Optional[ParsedPayload]
    response_payload: Optional[ParsedPayload]
    txn_id: str
    org_txn_id: str
    umn: str
    summary: str
    error_codes: str
    error_details: str


def parse_datetime_ist(value: str) -> str:
    if not value:
        return ""
    value = value.replace("Z", "+00:00")
    dt = datetime.fromisoformat(value)
    try:
        from zoneinfo import ZoneInfo

        return dt.astimezone(ZoneInfo(IST)).strftime("%Y-%m-%d %H:%M:%S %Z")
    except Exception:
        return dt.isoformat()


def strip_ns(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def sanitize_xml(xml: str) -> str:
    xml = re.sub(r"<Signature\b.*?</Signature>", "", xml, flags=re.S)
    xml = re.sub(r'\s+xmlns:_xmlns="xmlns"', "", xml)
    xml = re.sub(r"\s+_xmlns:[^=]+=\"[^\"]+\"", "", xml)
    xml = re.sub(r'\s+xmlns="http://www\.w3\.org/2000/09/xmldsig#"', "", xml)
    return xml.strip()


def flatten_xml(elem: ET.Element, path: Optional[str] = None) -> Dict[str, str]:
    tag = strip_ns(elem.tag)
    current = path or tag
    data: Dict[str, str] = {}
    for key, value in elem.attrib.items():
        data[f"{current}@{key}"] = value
    children = list(elem)
    text = (elem.text or "").strip()
    if text and not children:
        data[current] = text

    counts = Counter(strip_ns(child.tag) for child in children)
    seen: Dict[str, int] = defaultdict(int)
    for child in children:
        child_tag = strip_ns(child.tag)
        seen[child_tag] += 1
        suffix = f"[{seen[child_tag]}]" if counts[child_tag] > 1 else ""
        child_path = f"{current}.{child_tag}{suffix}"
        data.update(flatten_xml(child, child_path))
    return data


def normalize_flat_keys(flat: Dict[str, str], root_tag: str) -> Dict[str, str]:
    prefix = f"{root_tag}."
    normalized: Dict[str, str] = {}
    for key, value in flat.items():
        if key.startswith(prefix):
            normalized[key[len(prefix) :]] = value
        elif key == root_tag:
            normalized[""] = value
        else:
            normalized[key] = value
    return normalized


def parse_http_payload(value: str) -> Tuple[str, str]:
    value = value.replace("\r\n", "\n").strip()
    if "\n\n" in value:
        head, body = value.split("\n\n", 1)
        request_line = head.splitlines()[0] if head.splitlines() else ""
        return request_line, body.strip()
    return "", value


def parse_xml_body(body: str) -> Tuple[str, Dict[str, str], Dict[str, str]]:
    body = sanitize_xml(body)
    root = ET.fromstring(body)
    root_tag = strip_ns(root.tag)
    flat = flatten_xml(root, root_tag)
    return root_tag, flat, normalize_flat_keys(flat, root_tag)


def parse_json_message(raw: str) -> Dict[str, str]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    result = {}
    for key in ["eventID", "caller", "message", "api_target", "txn_id", "org_txn_id", "umn", "time"]:
        value = parsed.get(key)
        if value:
            result[key] = value
    return result


def unique_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    seen = set()
    deduped = []
    for row in rows:
        key = tuple(
            (row.get(column) or "")
            for column in [
                "Date",
                "api_flow",
                "api_dir",
                "_message",
                "message",
                "request_payload",
                "response_payload",
                "request",
                "response",
            ]
        )
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    return deduped


def build_parsed_payload(field_name: str, raw_value: str) -> Optional[ParsedPayload]:
    if not raw_value:
        return None
    request_line, body = parse_http_payload(raw_value)
    if not body.startswith("<"):
        return ParsedPayload(
            field_name=field_name,
            request_line=request_line,
            body=body,
            root_tag="",
            flat={},
            normalized_flat={},
        )
    root_tag, flat, normalized = parse_xml_body(body)
    return ParsedPayload(
        field_name=field_name,
        request_line=request_line,
        body=sanitize_xml(body),
        root_tag=root_tag,
        flat=flat,
        normalized_flat=normalized,
    )


def extract_ack_errors(flat: Dict[str, str]) -> Tuple[str, str]:
    codes = []
    details = []
    for key in sorted(flat):
        if key.endswith(".errorCd"):
            codes.append(flat[key])
        if key.endswith(".errorDtl"):
            details.append(flat[key])
    return ", ".join(codes), " | ".join(details)


def first_value(flat: Dict[str, str], *keys: str) -> str:
    for key in keys:
        if flat.get(key):
            return flat[key]
    return ""


def summarize_exchange(request: Optional[ParsedPayload], response: Optional[ParsedPayload], log_message: str) -> Tuple[str, str, str]:
    error_codes = ""
    error_details = ""
    parts = []

    if request and request.root_tag == "RespMandate":
        result = first_value(request.normalized_flat, "Resp@result")
        err_code = first_value(request.normalized_flat, "Resp@errCode")
        if result:
            parts.append(f"Incoming RespMandate {result}")
        if err_code:
            error_codes = err_code
            parts.append(f"errCode={err_code}")

    if request and request.root_tag == "RespChkTxn":
        result = first_value(request.normalized_flat, "Resp@result")
        err_code = first_value(request.normalized_flat, "Resp@errCode")
        if result:
            parts.append(f"Incoming RespChkTxn {result}")
        if err_code:
            error_codes = err_code
            parts.append(f"errCode={err_code}")

    if request and request.root_tag == "ReqMandateConfirmation":
        org_status = first_value(request.normalized_flat, "TxnConfirmation@orgStatus")
        org_err_code = first_value(request.normalized_flat, "TxnConfirmation@orgErrCode")
        if org_status:
            parts.append(f"Incoming ReqMandateConfirmation orgStatus={org_status}")
        if org_err_code:
            error_codes = org_err_code
            parts.append(f"orgErrCode={org_err_code}")

    if response and response.root_tag == "Ack":
        ack_api = first_value(response.normalized_flat, "Ack@api", "@api")
        ack_msg = first_value(response.normalized_flat, "Ack@reqMsgId", "@reqMsgId")
        ack_codes, ack_details = extract_ack_errors(response.flat)
        if ack_codes:
            error_codes = ", ".join(filter(None, [error_codes, ack_codes])).strip(", ")
            error_details = ack_details
            parts.append(f"Ack on {ack_api or 'request'} contains errors")
        else:
            parts.append(f"Ack received for {ack_api or 'request'}")
        if ack_msg:
            parts.append(f"reqMsgId={ack_msg}")

    if not parts and log_message:
        parts.append(log_message)

    return "; ".join(parts), error_codes, error_details


def classify_case(exchanges: List[Exchange]) -> str:
    roots = {
        payload.root_tag
        for exchange in exchanges
        for payload in [exchange.request_payload, exchange.response_payload]
        if payload and payload.root_tag
    }
    if "RespAuthMandate" in roots or "ReqAuthMandate" in roots:
        return "Intent-based mandate / authorization flow"
    if "ReqMandate" in roots or "RespMandate" in roots:
        return "Collect-based mandate flow"
    return "Mandate flow"


def primary_api(exchange: Exchange) -> str:
    request = exchange.request_payload
    response = exchange.response_payload
    if request and request.root_tag and request.root_tag != "Ack":
        return request.root_tag
    if response and response.root_tag:
        return response.root_tag
    return ""


def flow_stage_rank(case_type: str, exchange: Exchange) -> int:
    api = primary_api(exchange)
    if case_type.startswith("Intent"):
        order = {
            "ReqAuthMandate": 10,
            "RespAuthMandate": 20,
            "ReqMandateConfirmation": 30,
            "RespMandateConfirmation": 40,
        }
        return order.get(api, 99)
    if case_type.startswith("Collect"):
        order = {
            "ReqMandate": 10,
            "ReqChkTxn": 20,
            "RespChkTxn": 30,
            "RespMandate": 40,
        }
        return order.get(api, 99)
    return 99


def case_error_codes(exchanges: List[Exchange]) -> List[str]:
    codes = []
    for exchange in exchanges:
        if not exchange.error_codes:
            continue
        for code in [item.strip() for item in exchange.error_codes.split(",")]:
            if code and code not in codes:
                codes.append(code)
    return codes


def find_exchange(exchanges: List[Exchange], api_name: str) -> Optional[Exchange]:
    for exchange in exchanges:
        if primary_api(exchange) == api_name:
            return exchange
    return None


def describe_case_outcome(case_type: str, exchanges: List[Exchange]) -> str:
    intent_auth = find_exchange(exchanges, "RespAuthMandate")
    intent_conf = find_exchange(exchanges, "ReqMandateConfirmation")
    collect_req = find_exchange(exchanges, "ReqMandate")
    collect_resp = find_exchange(exchanges, "RespMandate")
    collect_chk = find_exchange(exchanges, "RespChkTxn")

    if case_type.startswith("Intent"):
        if intent_auth and not intent_auth.error_codes and intent_conf and intent_conf.error_codes:
            return (
                "ReqAuthMandate and RespAuthMandate were acknowledged successfully. "
                f"NPCI then sent ReqMandateConfirmation with {intent_conf.summary}."
            )
        if intent_auth and intent_auth.error_codes:
            return f"RespAuthMandate returned Ack errors: {intent_auth.error_codes}."
        return "Intent-based mandate flow observed."

    if case_type.startswith("Collect"):
        if collect_req and not collect_req.error_codes and collect_resp and collect_resp.error_codes:
            outcome = (
                "ReqMandate was acknowledged successfully. "
                f"NPCI then sent RespMandate with {collect_resp.summary}."
            )
            if not collect_chk:
                outcome += " No ReqChkTxn/RespChkTxn exchange is present in the attached collect-flow log."
            return outcome
        if collect_chk and collect_chk.error_codes:
            return f"Collect flow shows RespChkTxn failure: {collect_chk.summary}."
        return "Collect-based mandate flow observed."

    return "Mandate flow observed."


def build_npci_questions(cases: List[Tuple[str, str, List[Exchange], int]]) -> List[str]:
    all_codes = []
    for _, _, exchanges, _ in cases:
        all_codes.extend(case_error_codes(exchanges))
    unique_codes = []
    for code in all_codes:
        if code not in unique_codes:
            unique_codes.append(code)

    if unique_codes == ["U16"] or (unique_codes and all(code == "U16" for code in unique_codes)):
        return [
            "Post 2.17 enablement, both attached flows now fail with U16. Please confirm the exact NPCI meaning and risk trigger for U16 in these transactions.",
            "In the intent flow, ReqAuthMandate and RespAuthMandate were acknowledged successfully. Why was ReqMandateConfirmation raised with orgStatus=FAILURE and orgErrCode=U16?",
            "In the collect flow, ReqMandate was acknowledged successfully. Why did RespMandate return result=FAILURE with errCode=U16?",
            "Please confirm whether U16 is being triggered by NPCI-side risk checks, payer-bank risk controls, or another ecosystem participant.",
        ]

    return [
        "Please confirm the exact NPCI failure reason for the attached transaction IDs and indicate the stage where the flow is being rejected.",
        "Please confirm whether the failure is originating at NPCI, payer bank, payer app, or another ecosystem participant.",
        "Please review the XML evidence and confirm whether any payload field is non-compliant for the attached requests.",
        "Please share the NPCI-side interpretation of the observed error codes for these transactions.",
    ]


def extract_exchange_from_row(
    case_name: str,
    source_file: str,
    row: Dict[str, str],
    sequence_index: int,
) -> Optional[Exchange]:
    message_meta = parse_json_message(row.get("_message") or "")
    api_target = message_meta.get("api_target", "")
    caller = message_meta.get("caller", "")
    log_message = message_meta.get("message", "") or (row.get("message") or "")

    request_raw = row.get("request_payload") or row.get("request") or ""
    response_raw = row.get("response_payload") or row.get("response") or ""
    request_payload = build_parsed_payload("request", request_raw) if request_raw else None
    response_payload = build_parsed_payload("response", response_raw) if response_raw else None

    has_upi_xml = any(
        payload and payload.root_tag
        for payload in [request_payload, response_payload]
    )
    if not has_upi_xml:
        return None

    txn_id = (
        row.get("api_target")
        or message_meta.get("txn_id", "")
        or first_value(
            (request_payload.normalized_flat if request_payload else {}),
            "Txn@id",
            "Mandate@txnId",
        )
        or row.get("api_target", "")
    )
    txn_id = first_value(
        request_payload.normalized_flat if request_payload else {},
        "Txn@id",
        "Mandate@txnId",
    ) or first_value(
        response_payload.normalized_flat if response_payload else {},
        "Txn@id",
        "Mandate@txnId",
    ) or message_meta.get("txn_id", "")

    org_txn_id = first_value(
        request_payload.normalized_flat if request_payload else {},
        "Txn@orgTxnId",
    ) or first_value(
        response_payload.normalized_flat if response_payload else {},
        "Txn@orgTxnId",
    ) or message_meta.get("org_txn_id", "")

    umn = first_value(
        request_payload.normalized_flat if request_payload else {},
        "Mandate@umn",
    ) or first_value(
        response_payload.normalized_flat if response_payload else {},
        "Mandate@umn",
    ) or message_meta.get("umn", "")

    summary, error_codes, error_details = summarize_exchange(request_payload, response_payload, log_message)

    return Exchange(
        sequence_index=sequence_index,
        case_name=case_name,
        case_type="",
        source_file=source_file,
        timestamp_utc=row.get("Date", ""),
        timestamp_ist=parse_datetime_ist(row.get("Date", "")),
        api_flow=row.get("api_flow", ""),
        api_dir=row.get("api_dir", ""),
        api_target=api_target,
        caller=caller,
        level=row.get("log_level") or row.get("level") or "",
        log_message=log_message,
        request_payload=request_payload,
        response_payload=response_payload,
        txn_id=txn_id,
        org_txn_id=org_txn_id,
        umn=umn,
        summary=summary,
        error_codes=error_codes,
        error_details=error_details,
    )


def read_case(path: Path) -> Tuple[List[Exchange], int]:
    with path.open(newline="") as handle:
        rows = list(csv.DictReader(handle))
    deduped = unique_rows(rows)
    case_name = path.stem
    exchanges = []
    for index, row in enumerate(deduped):
        exchange = extract_exchange_from_row(case_name, str(path), row, index)
        if exchange:
            exchanges.append(exchange)
    case_type = classify_case(exchanges)
    for exchange in exchanges:
        exchange.case_type = case_type
    exchanges.sort(
        key=lambda item: (
            item.timestamp_utc,
            flow_stage_rank(case_type, item),
            item.sequence_index,
        )
    )
    return exchanges, len(deduped)


def outgoing_payload(exchanges: List[Exchange], root_tag: str) -> Optional[ParsedPayload]:
    for exchange in exchanges:
        payload = exchange.request_payload
        if payload and payload.root_tag == root_tag:
            return payload
    return None


def auto_fit(ws) -> None:
    max_widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = min(max(len(str(cell.value)), 0), 120)
            max_widths[cell.column] = max(max_widths.get(cell.column, 0), length)
    for column, width in max_widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(width + 2, 60)


def style_table_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_cells(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_summary_sheet(
    wb: Workbook,
    cases: List[Tuple[str, str, List[Exchange], int]],
    diff_rows: List[List[str]],
    notes: Optional[List[str]] = None,
) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "SBMD NPCI Log Analysis"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Prepared On"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Notes"
    base_note = "Report is based on deduplicated CSV log exports. Timestamps shown in IST."
    if notes:
        base_note = " ".join([base_note] + notes)
    ws["B4"] = base_note

    start_row = 6
    headers = [
        "Case",
        "Flow Type",
        "Primary Txn ID",
        "Observed Outcome",
        "Error Codes",
        "Source File",
        "Deduplicated Events",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(start_row, col, header)
    style_table_header(ws[start_row])

    row_index = start_row + 1
    for case_name, case_type, exchanges, deduped_count in cases:
        txn_ids = [item.txn_id for item in exchanges if item.txn_id]
        primary_txn_id = txn_ids[0] if txn_ids else ""
        error_codes = ", ".join(sorted({item.error_codes for item in exchanges if item.error_codes}))
        outcome = describe_case_outcome(case_type, exchanges)
        ws.cell(row_index, 1, case_name)
        ws.cell(row_index, 2, case_type)
        ws.cell(row_index, 3, primary_txn_id)
        ws.cell(row_index, 4, outcome)
        ws.cell(row_index, 5, error_codes)
        ws.cell(row_index, 6, exchanges[0].source_file if exchanges else "")
        ws.cell(row_index, 7, deduped_count)
        row_index += 1

    row_index += 1
    ws.cell(row_index, 1, "NPCI-Facing Questions")
    ws.cell(row_index, 1).font = Font(bold=True)
    ws.cell(row_index, 1).fill = SUBHEADER_FILL
    row_index += 1
    questions = build_npci_questions(cases)
    for question in questions:
        ws.cell(row_index, 1, question)
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        row_index += 1

    row_index += 1
    ws.cell(row_index, 1, "Payload Comparison Note")
    ws.cell(row_index, 1).font = Font(bold=True)
    ws.cell(row_index, 1).fill = SUBHEADER_FILL
    row_index += 1
    note = (
        "Business fields across intent-flow RespAuthMandate and collect-flow ReqMandate are largely aligned for amount, recurrence, "
        "validity, payer/payee addressing, and mandate type. Flow-specific differences are listed in the Payload_Comparison sheet."
    )
    ws.cell(row_index, 1, note)
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)

    style_cells(ws)
    auto_fit(ws)


def add_case_sheet(wb: Workbook, title: str, exchanges: List[Exchange]) -> None:
    ws = wb.create_sheet(title)
    headers = [
        "Seq",
        "Timestamp IST",
        "Direction",
        "API Flow",
        "Primary API",
        "Txn ID",
        "Org Txn ID",
        "Message ID",
        "Req Msg ID",
        "Summary",
        "Error Codes",
        "Error Details",
    ]
    ws.append(headers)
    style_table_header(ws[1])

    for index, exchange in enumerate(exchanges, start=1):
        request = exchange.request_payload
        response = exchange.response_payload
        primary_api_name = primary_api(exchange)

        msg_id = first_value(
            request.normalized_flat if request else {},
            "Head@msgId",
        ) or first_value(
            response.normalized_flat if response else {},
            "Head@msgId",
        )
        req_msg_id = first_value(
            request.normalized_flat if request else {},
            "Resp@reqMsgId",
            "Ack@reqMsgId",
            "@reqMsgId",
        ) or first_value(
            response.normalized_flat if response else {},
            "Resp@reqMsgId",
            "Ack@reqMsgId",
            "@reqMsgId",
        )

        ws.append(
            [
                index,
                exchange.timestamp_ist,
                exchange.api_dir or "",
                exchange.api_flow,
                primary_api_name,
                exchange.txn_id,
                exchange.org_txn_id,
                msg_id,
                req_msg_id,
                exchange.summary,
                exchange.error_codes,
                exchange.error_details,
            ]
        )

    style_cells(ws)
    ws.freeze_panes = "A2"
    auto_fit(ws)

    for row in ws.iter_rows(min_row=2):
        error_value = str(row[10].value or "")
        if error_value:
            for cell in row:
                cell.fill = ERROR_FILL


def add_payload_comparison_sheet(
    wb: Workbook,
    intent_payload: Optional[ParsedPayload],
    collect_payload: Optional[ParsedPayload],
) -> None:
    ws = wb.create_sheet("Payload_Comparison")
    headers = [
        "Field",
        "Intent Flow RespAuthMandate",
        "Collect Flow ReqMandate",
        "Comparison",
    ]
    ws.append(headers)
    style_table_header(ws[1])

    comparison_fields = [
        "Txn@type",
        "Txn@purpose",
        "Txn@initiatedBy",
        "Txn@initiationMode",
        "Txn@refCategory",
        "Txn@refUrl",
        "Mandate@blockFund",
        "Mandate@revokeable",
        "Mandate@shareToPayee",
        "Mandate@name",
        "Mandate.Amount@rule",
        "Mandate.Amount@value",
        "Mandate.Recurrence@pattern",
        "Mandate.Validity@start",
        "Mandate.Validity@end",
        "Mandate@umn",
        "Payer@addr",
        "Payer@code",
        "Payer@type",
        "Payees.Payee@addr",
        "Payees.Payee@code",
        "Payees.Payee@type",
    ]

    for field in comparison_fields:
        left = intent_payload.normalized_flat.get(field, "") if intent_payload else ""
        right = collect_payload.normalized_flat.get(field, "") if collect_payload else ""
        if left == right and left:
            status = "Same"
        elif left and right:
            status = "Different"
        elif left or right:
            status = "Present in one flow only"
        else:
            status = ""
        ws.append([field, left, right, status])

    style_cells(ws)
    ws.freeze_panes = "A2"
    auto_fit(ws)

    for row in ws.iter_rows(min_row=2):
        status = row[3].value
        fill = GOOD_FILL if status == "Same" else HIGHLIGHT_FILL
        for cell in row:
            cell.fill = fill


def add_xml_evidence_sheet(wb: Workbook, cases: List[Tuple[str, str, List[Exchange], int]]) -> None:
    ws = wb.create_sheet("XML_Evidence")
    headers = [
        "Case",
        "Timestamp IST",
        "Primary API",
        "API Flow",
        "Request Line",
        "Response Line",
        "Request XML",
        "Response XML",
    ]
    ws.append(headers)
    style_table_header(ws[1])

    for case_name, _, exchanges, _ in cases:
        for exchange in exchanges:
            request = exchange.request_payload
            response = exchange.response_payload
            ws.append(
                [
                    case_name,
                    exchange.timestamp_ist,
                    request.root_tag if request else "",
                    exchange.api_flow,
                    request.request_line if request else "",
                    response.request_line if response else "",
                    request.body if request else "",
                    response.body if response else "",
                ]
            )

    style_cells(ws)
    ws.freeze_panes = "A2"
    auto_fit(ws)
    ws.column_dimensions["G"].width = 80
    ws.column_dimensions["H"].width = 80


def add_raw_events_sheet(wb: Workbook, cases: List[Tuple[str, str, List[Exchange], int]]) -> None:
    ws = wb.create_sheet("Raw_Events")
    headers = [
        "Case",
        "Timestamp IST",
        "API Flow",
        "Direction",
        "Caller",
        "Log Level",
        "Log Message",
        "Txn ID",
        "Org Txn ID",
        "UMN",
        "Summary",
    ]
    ws.append(headers)
    style_table_header(ws[1])

    for case_name, _, exchanges, _ in cases:
        for exchange in exchanges:
            ws.append(
                [
                    case_name,
                    exchange.timestamp_ist,
                    exchange.api_flow,
                    exchange.api_dir,
                    exchange.caller,
                    exchange.level,
                    exchange.log_message,
                    exchange.txn_id,
                    exchange.org_txn_id,
                    exchange.umn,
                    exchange.summary,
                ]
            )

    style_cells(ws)
    ws.freeze_panes = "A2"
    auto_fit(ws)


def build_report(output_path: Path, input_paths: List[Path], notes: Optional[List[str]] = None) -> None:
    case_records: List[Tuple[str, str, List[Exchange], int]] = []
    for path in input_paths:
        exchanges, deduped_count = read_case(path)
        case_type = exchanges[0].case_type if exchanges else "Mandate flow"
        case_records.append((path.stem, case_type, exchanges, deduped_count))

    wb = Workbook()

    intent_exchanges = next((records for _, case_type, records, _ in case_records if case_type.startswith("Intent")), [])
    collect_exchanges = next((records for _, case_type, records, _ in case_records if case_type.startswith("Collect")), [])
    intent_payload = outgoing_payload(intent_exchanges, "RespAuthMandate")
    collect_payload = outgoing_payload(collect_exchanges, "ReqMandate")

    add_summary_sheet(wb, case_records, [], notes=notes)
    for case_name, case_type, exchanges, _ in case_records:
        if case_type.startswith("Intent"):
            title = "Intent_Flow"
        elif case_type.startswith("Collect"):
            title = "Collect_Flow"
        else:
            title = re.sub(r"[^A-Za-z0-9_]", "_", case_name)[:31]
        add_case_sheet(wb, title, exchanges)
    add_payload_comparison_sheet(wb, intent_payload, collect_payload)
    add_xml_evidence_sheet(wb, case_records)
    add_raw_events_sheet(wb, case_records)

    wb.save(output_path)


def main(argv: List[str]) -> int:
    if len(argv) < 3:
        print("Usage: sbmd_npci_report.py <output.xlsx> <input1.csv> <input2.csv> [more.csv...]", file=sys.stderr)
        return 1

    output_path = Path(argv[1]).expanduser().resolve()
    input_paths = [Path(item).expanduser().resolve() for item in argv[2:]]
    build_report(output_path, input_paths)
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
